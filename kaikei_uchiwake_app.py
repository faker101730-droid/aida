
import io
from pathlib import Path
import pandas as pd
import streamlit as st

st.set_page_config(page_title="会計内訳アプリ Pro ", page_icon="📘", layout="wide")

st.markdown("""
<style>
.block-container {padding-top: 1.0rem; padding-bottom: 2rem;}
.kpi-card {
    background: linear-gradient(135deg, #0f172a 0%, #111827 100%);
    border: 1px solid #334155;
    border-radius: 18px;
    padding: 18px 18px 14px 18px;
    min-height: 118px;
    box-shadow: 0 8px 20px rgba(0,0,0,0.18);
}
.kpi-label {font-size: 0.9rem; color: #cbd5e1; margin-bottom: 6px;}
.kpi-value {font-size: 2rem; font-weight: 700; color: white; line-height: 1.1;}
.kpi-note {font-size: 0.8rem; color: #94a3b8; margin-top: 6px;}
.info-box {
    background: #0b3b2f;
    border: 1px solid #14532d;
    border-radius: 14px;
    padding: 12px 14px;
    color: #dcfce7;
    margin-bottom: 0.5rem;
}
.subtle-box {
    background: #0f172a;
    border: 1px solid #1e293b;
    border-radius: 14px;
    padding: 12px 14px;
    color: #cbd5e1;
    margin-bottom: 0.5rem;
}
</style>
""", unsafe_allow_html=True)

INIT_REQUIRED = ["基準日", "勘定科目", "相手先", "初期残高"]
STANDARD_REQUIRED = ["日付", "借方科目", "貸方科目", "金額", "相手先", "摘要"]

COLUMN_ALIASES = {
    "日付": ["日付", "伝票日付", "起票日", "処理日"],
    "借方中区分": ["借方中区分", "借方科目（中区分）", "借方科目(中区分)", "借方中分類", "借方中科目"],
    "借方小区分": ["借方小区分", "借方科目（小区分）", "借方科目(小区分)", "借方小分類", "借方小科目"],
    "借方補助区分": ["借方補助区分", "借方補助", "借方補助科目", "借方補助区分名"],
    "借方取引先": ["借方取引先", "借方先方", "借方相手先", "借方補助先", "借方取引先名"],
    "借方金額": ["借方金額", "借方金額（円）", "借方金額(円)"],
    "貸方中区分": ["貸方中区分", "貸方科目（中区分）", "貸方科目(中区分)", "貸方中分類", "貸方中科目"],
    "貸方小区分": ["貸方小区分", "貸方科目（小区分）", "貸方科目(小区分)", "貸方小分類", "貸方小科目"],
    "貸方補助区分": ["貸方補助区分", "貸方補助", "貸方補助科目", "貸方補助区分名"],
    "貸方取引先": ["貸方取引先", "貸方先方", "貸方相手先", "貸方補助先", "貸方取引先名"],
    "貸方金額": ["貸方金額", "貸方金額（円）", "貸方金額(円)"],
    "摘要文": ["摘要文", "摘要", "摘要内容", "取引摘要", "明細摘要"],
    "伝票番号": ["伝票番号", "伝票No", "伝票NO", "伝票Ｎｏ", "伝票番号No"],
}

BS_DEFAULT = [
    "現金","現金預金","普通預金","小口現金","立替金","未収金","未収入金","前払費用","前払金","仮払金","貸付金",
    "未払金","未払費用","預り金","前受金","仮受金","借入金","賞与引当金","退職給付引当金","繰越利益剰余金"
]

def normalize_name(x: str) -> str:
    return str(x).replace("\n", "").replace("\r", "").replace(" ", "").replace("　", "").strip()

def normalize_columns(df: pd.DataFrame) -> pd.DataFrame:
    df = df.copy()
    df.columns = [normalize_name(c) for c in df.columns]
    return df

def read_any_table(uploaded_file, header=None):
    suffix = Path(uploaded_file.name).suffix.lower()
    if suffix in [".xlsx", ".xlsm", ".xls"]:
        uploaded_file.seek(0)
        return pd.read_excel(uploaded_file, header=header)
    for enc in ["utf-8-sig", "cp932", "utf-8", "shift_jis"]:
        try:
            uploaded_file.seek(0)
            return pd.read_csv(uploaded_file, header=header, encoding=enc)
        except Exception:
            continue
    uploaded_file.seek(0)
    return pd.read_csv(uploaded_file, header=header)

def detect_header_row(raw_df: pd.DataFrame, candidate_cols):
    candidate_cols = set([normalize_name(x) for x in candidate_cols])
    max_scan = min(len(raw_df), 10)
    for i in range(max_scan):
        vals = set(normalize_name(v) for v in raw_df.iloc[i].tolist())
        if candidate_cols.issubset(vals):
            return i
    return 0

def alias_exists(colset, canonical):
    aliases = [normalize_name(x) for x in COLUMN_ALIASES.get(canonical, [canonical])]
    return any(a in colset for a in aliases)

def find_alias(colset, canonical):
    aliases = [normalize_name(x) for x in COLUMN_ALIASES.get(canonical, [canonical])]
    for a in aliases:
        if a in colset:
            return a
    return None

def detect_file_kind(df: pd.DataFrame):
    cols = set(normalize_columns(df).columns)
    if set(INIT_REQUIRED).issubset(cols):
        return "initial"
    if set(STANDARD_REQUIRED).issubset(cols):
        return "journal_standard"
    fukushi_core = ["日付", "借方中区分", "貸方中区分", "借方金額", "貸方金額", "摘要文"]
    if all(alias_exists(cols, c) for c in fukushi_core):
        return "journal_fukushi"
    return "unknown"

def parse_file(uploaded_file):
    raw = read_any_table(uploaded_file, header=None)
    init_row = detect_header_row(raw, INIT_REQUIRED)
    standard_row = detect_header_row(raw, STANDARD_REQUIRED)
    fukushi_row = detect_header_row(raw, [COLUMN_ALIASES[k][0] for k in ["日付","借方中区分","貸方中区分","借方金額","貸方金額","摘要文"]])

    candidates = []
    for row in [init_row, standard_row, fukushi_row]:
        try:
            df = normalize_columns(read_any_table(uploaded_file, header=row))
            candidates.append(df)
        except Exception:
            pass

    for df in candidates:
        if detect_file_kind(df) == "initial":
            return "initial", df
    for df in candidates:
        if detect_file_kind(df) in ["journal_standard", "journal_fukushi"]:
            return "journal", df

    raise ValueError("必要列を判定できませんでした。初期残高か仕訳マスタの形式を確認してください。")

def to_num(series):
    return pd.to_numeric(
        series.astype(str)
        .str.replace(",", "", regex=False)
        .str.replace("¥", "", regex=False)
        .str.replace("円", "", regex=False)
        .str.replace("△", "-", regex=False)
        .str.strip(),
        errors="coerce"
    )

def prepare_initial(df):
    df = normalize_columns(df)
    missing = [c for c in INIT_REQUIRED if c not in df.columns]
    if missing:
        raise ValueError("初期残高に必要列が不足しています: " + ", ".join(missing))
    out = df.copy()
    out["基準日"] = pd.to_datetime(out["基準日"], errors="coerce")
    out["初期残高"] = to_num(out["初期残高"])
    out["勘定科目"] = out["勘定科目"].fillna("").astype(str).str.strip()
    out["相手先"] = out["相手先"].fillna("(空欄)").astype(str).str.strip().replace("", "(空欄)")
    if out["基準日"].isna().any():
        raise ValueError("初期残高の基準日に日付変換できない値があります。")
    if out["初期残高"].isna().any():
        raise ValueError("初期残高に数値変換できない値があります。")
    return out

def clean_account_name(s):
    if pd.isna(s):
        return ""
    s = str(s).strip()
    s = pd.Series([s]).str.replace(r"^\([^\)]*\)", "", regex=True).iloc[0].strip()
    return s


def make_display_counterparty(primary_partner, small_label, sub_label, note):
    primary = "" if pd.isna(primary_partner) else str(primary_partner).strip()
    small = "" if pd.isna(small_label) else str(small_label).strip()
    sub = "" if pd.isna(sub_label) else str(sub_label).strip()
    note = "" if pd.isna(note) else str(note).strip()

    if primary:
        return primary
    if sub:
        return f"補助:{sub}"
    if small:
        return f"小区分:{small}"
    if note:
        short_note = note[:24] + ("…" if len(note) > 24 else "")
        return f"摘要:{short_note}"
    return "(空欄)"

def choose_partner(primary, secondary):
    p = "" if pd.isna(primary) else str(primary).strip()
    s = "" if pd.isna(secondary) else str(secondary).strip()
    return p if p else s

def standardize_fukushi_columns(df):
    df = normalize_columns(df).copy()
    colset = set(df.columns)
    rename_map = {}
    for canonical in COLUMN_ALIASES.keys():
        found = find_alias(colset, canonical)
        if found and found != canonical:
            rename_map[found] = canonical
    if rename_map:
        df = df.rename(columns=rename_map)
    return df

def prepare_journal(df, grain_mode="中区分"):
    df = normalize_columns(df)
    kind = detect_file_kind(df)

    if kind == "journal_standard":
        out = df.copy()
        out["日付"] = pd.to_datetime(out["日付"], errors="coerce")
        out["金額"] = to_num(out["金額"])
        for c in ["借方科目", "貸方科目", "相手先", "摘要"]:
            out[c] = out[c].fillna("").astype(str).str.strip()
        out["相手先_借方"] = out["相手先"].replace("", "(空欄)")
        out["相手先_貸方"] = out["相手先"].replace("", "(空欄)")
        if "伝票番号" not in out.columns:
            out["伝票番号"] = ""
        out["借方中区分表示"] = out["借方科目"]
        out["貸方中区分表示"] = out["貸方科目"]
        out["借方小区分表示"] = out["借方科目"]
        out["貸方小区分表示"] = out["貸方科目"]
        out["借方補助区分"] = out.get("借方補助区分", "")
        out["貸方補助区分"] = out.get("貸方補助区分", "")
        if out["日付"].isna().any():
            raise ValueError("仕訳データの日付に日付変換できない値があります。")
        if out["金額"].isna().any():
            raise ValueError("仕訳データの金額に数値変換できない値があります。")
        return out

    if kind == "journal_fukushi":
        out = standardize_fukushi_columns(df)
        required = ["日付", "借方中区分", "貸方中区分", "借方金額", "貸方金額", "摘要文"]
        missing = [c for c in required if c not in out.columns]
        if missing:
            raise ValueError("福祉の森データに必要列が不足しています: " + ", ".join(missing))

        out["日付"] = pd.to_datetime(out["日付"], errors="coerce")
        out["借方金額"] = to_num(out["借方金額"])
        out["貸方金額"] = to_num(out["貸方金額"])
        out["金額"] = out["借方金額"].where(out["借方金額"].notna(), out["貸方金額"])

        for c in ["借方小区分","借方補助区分","借方取引先","貸方小区分","貸方補助区分","貸方取引先","伝票番号"]:
            if c not in out.columns:
                out[c] = ""

        out["借方中区分表示"] = out["借方中区分"].fillna("").map(clean_account_name)
        out["貸方中区分表示"] = out["貸方中区分"].fillna("").map(clean_account_name)
        out["借方小区分表示"] = out["借方小区分"].fillna("").map(clean_account_name)
        out["貸方小区分表示"] = out["貸方小区分"].fillna("").map(clean_account_name)

        if grain_mode == "小区分":
            out["借方科目"] = out["借方小区分表示"].where(out["借方小区分表示"] != "", out["借方中区分表示"])
            out["貸方科目"] = out["貸方小区分表示"].where(out["貸方小区分表示"] != "", out["貸方中区分表示"])
        else:
            out["借方科目"] = out["借方中区分表示"]
            out["貸方科目"] = out["貸方中区分表示"]

        out["摘要"] = out["摘要文"].fillna("").astype(str).str.strip()
        out["相手先_借方"] = [
            make_display_counterparty(a, s, b, n)
            for a, s, b, n in zip(out["借方取引先"], out["借方小区分表示"], out["借方補助区分"], out["摘要"])
        ]
        out["相手先_貸方"] = [
            make_display_counterparty(a, s, b, n)
            for a, s, b, n in zip(out["貸方取引先"], out["貸方小区分表示"], out["貸方補助区分"], out["摘要"])
        ]
        out["相手先"] = out["相手先_借方"]

        if out["日付"].isna().any():
            raise ValueError("福祉の森データの日付に日付変換できない値があります。")
        if out["金額"].isna().any():
            raise ValueError("福祉の森データの金額に数値変換できない値があります。")

        return out

    raise ValueError("仕訳データの形式を認識できませんでした。")

def choose_accounts(initial_df, journal_df, manual_text):
    manual = [x.strip() for x in manual_text.splitlines() if x.strip()]
    all_accounts = sorted(
        {str(x).strip() for x in initial_df["勘定科目"].dropna().tolist()}
        | {str(x).strip() for x in journal_df["借方科目"].dropna().tolist()}
        | {str(x).strip() for x in journal_df["貸方科目"].dropna().tolist()}
    )
    all_accounts = [x for x in all_accounts if x]
    if manual:
        filtered = [a for a in all_accounts if a in manual]
        return filtered if filtered else all_accounts
    return all_accounts

def month_options(initial_df, journal_df):
    base = initial_df["基準日"].max()
    return sorted(journal_df[journal_df["日付"] > base]["日付"].dt.to_period("M").astype(str).unique().tolist())

def make_period_labels(display_mode, start_month, end_month):
    if display_mode == "単月" or start_month == end_month:
        return start_month, f"対象月: {start_month}"
    return f"{start_month}_to_{end_month}", f"対象期間: {start_month} ～ {end_month}"

def build_summary(initial_df, journal_df, account, start_month, end_month=None, mode="単月"):
    if end_month is None:
        end_month = start_month

    start_period = pd.Period(start_month)
    end_period = pd.Period(end_month)
    period_start = start_period.start_time
    period_end = end_period.end_time
    init_base = initial_df["基準日"].max()

    if period_start <= init_base:
        raise ValueError("開始月が初期残高基準日以前です。基準日より後の月を選択してください。")
    if end_period < start_period:
        raise ValueError("終了月は開始月以降を選択してください。")

    init_acc = initial_df[initial_df["勘定科目"] == account].copy()
    before = journal_df[(journal_df["日付"] > init_base) & (journal_df["日付"] < period_start)].copy()
    current = journal_df[(journal_df["日付"] >= period_start) & (journal_df["日付"] <= period_end)].copy()

    init_partner = init_acc.groupby("相手先", dropna=False)["初期残高"].sum().rename("期首初期残高").reset_index()
    inc_before = before[before["借方科目"] == account].groupby("相手先_借方", dropna=False)["金額"].sum().rename("前期間まで増加").reset_index().rename(columns={"相手先_借方": "相手先"})
    dec_before = before[before["貸方科目"] == account].groupby("相手先_貸方", dropna=False)["金額"].sum().rename("前期間まで減少").reset_index().rename(columns={"相手先_貸方": "相手先"})
    inc_cur = current[current["借方科目"] == account].groupby("相手先_借方", dropna=False)["金額"].sum().rename("対象期間増加").reset_index().rename(columns={"相手先_借方": "相手先"})
    dec_cur = current[current["貸方科目"] == account].groupby("相手先_貸方", dropna=False)["金額"].sum().rename("対象期間減少").reset_index().rename(columns={"相手先_貸方": "相手先"})

    summary = init_partner.merge(inc_before, on="相手先", how="outer") \
                         .merge(dec_before, on="相手先", how="outer") \
                         .merge(inc_cur, on="相手先", how="outer") \
                         .merge(dec_cur, on="相手先", how="outer")
    if summary.empty:
        summary = pd.DataFrame({"相手先": []})

    for col in ["期首初期残高", "前期間まで増加", "前期間まで減少", "対象期間増加", "対象期間減少"]:
        if col not in summary.columns:
            summary[col] = 0
        summary[col] = pd.to_numeric(summary[col], errors="coerce").fillna(0)

    summary["期首残高"] = summary["期首初期残高"] + summary["前期間まで増加"] - summary["前期間まで減少"]
    summary["期末残高"] = summary["期首残高"] + summary["対象期間増加"] - summary["対象期間減少"]
    summary["対象期間動きなし"] = (summary["対象期間増加"].abs() + summary["対象期間減少"].abs()) == 0
    summary["マイナス残高"] = summary["期末残高"] < 0

    movement = pd.concat([
        before[before["借方科目"] == account][["日付", "相手先_借方"]].rename(columns={"相手先_借方": "相手先"}),
        before[before["貸方科目"] == account][["日付", "相手先_貸方"]].rename(columns={"相手先_貸方": "相手先"}),
        current[current["借方科目"] == account][["日付", "相手先_借方"]].rename(columns={"相手先_借方": "相手先"}),
        current[current["貸方科目"] == account][["日付", "相手先_貸方"]].rename(columns={"相手先_貸方": "相手先"}),
    ], ignore_index=True)

    if not movement.empty:
        last_move = movement.groupby("相手先", dropna=False)["日付"].max().reset_index().rename(columns={"日付": "最終更新日"})
        summary = summary.merge(last_move, on="相手先", how="left")
        summary["最終更新日"] = pd.to_datetime(summary["最終更新日"], errors="coerce")
        summary["経過日数"] = (period_end.normalize() - summary["最終更新日"]).dt.days
        summary["長期残存候補"] = (summary["期末残高"] != 0) & (summary["経過日数"] >= 60)
    else:
        summary["最終更新日"] = pd.NaT
        summary["経過日数"] = None
        summary["長期残存候補"] = False

    latest_note = current[(current["借方科目"] == account) | (current["貸方科目"] == account)].copy()
    if not latest_note.empty:
        latest_note["表示相手先"] = latest_note["相手先_借方"]
        latest_note.loc[latest_note["貸方科目"] == account, "表示相手先"] = latest_note.loc[latest_note["貸方科目"] == account, "相手先_貸方"]
        latest_note = latest_note.sort_values("日付").groupby("表示相手先", dropna=False)["摘要"].last().reset_index().rename(columns={"表示相手先":"相手先","摘要":"最新摘要"})
        latest_note["最新摘要"] = latest_note["最新摘要"].fillna("").astype(str)
        summary = summary.merge(latest_note, on="相手先", how="left")
    else:
        summary["最新摘要"] = ""
    summary["最新摘要"] = summary["最新摘要"].fillna("").replace("None", "")

    summary = summary[["相手先","最新摘要","最終更新日","期首残高","対象期間増加","対象期間減少","期末残高","長期残存候補","マイナス残高"]].copy()
    summary = summary.sort_values(["期末残高","相手先"], ascending=[False, True]).reset_index(drop=True)

    history_detail = before[(before["借方科目"] == account) | (before["貸方科目"] == account)].copy()
    current_detail = current[(current["借方科目"] == account) | (current["貸方科目"] == account)].copy()

    for d in [history_detail, current_detail]:
        if d.empty:
            continue
        d["表示相手先"] = d["相手先_借方"]
        d["対象側"] = "借方"
        d.loc[d["貸方科目"] == account, "表示相手先"] = d.loc[d["貸方科目"] == account, "相手先_貸方"]
        d.loc[d["貸方科目"] == account, "対象側"] = "貸方"

    return summary, history_detail, current_detail, init_base, period_start, period_end

def fmt_yen(x):
    try:
        return f"{int(round(float(x), 0)):,}"
    except Exception:
        return x

def card(label, value, note):
    st.markdown(f"""
    <div class="kpi-card">
        <div class="kpi-label">{label}</div>
        <div class="kpi-value">{fmt_yen(value)}</div>
        <div class="kpi-note">{note}</div>
    </div>
    """, unsafe_allow_html=True)

def style_summary(df):
    show = df.copy()
    if "最新摘要" in show.columns:
        show["最新摘要"] = show["最新摘要"].fillna("").replace("None", "")
    if "最終更新日" in show.columns:
        show["最終更新日"] = pd.to_datetime(show["最終更新日"], errors="coerce").dt.strftime("%Y-%m-%d").fillna("")
    for c in ["期首残高","対象期間増加","対象期間減少","期末残高"]:
        if c in show.columns:
            show[c] = show[c].apply(fmt_yen)
    show["長期残存候補"] = show["長期残存候補"].map(lambda x: "⚠" if x else "")
    show["マイナス残高"] = show["マイナス残高"].map(lambda x: "⚠" if x else "")
    return show.rename(columns={"対象期間増加":"増加","対象期間減少":"減少","長期残存候補":"長期残存","マイナス残高":"マイナス"})

def style_detail(df):
    cols = [c for c in ["日付","対象側","表示相手先","借方科目","貸方科目","借方中区分表示","借方小区分表示","貸方中区分表示","貸方小区分表示","金額","摘要","伝票番号"] if c in df.columns]
    show = df[cols].copy() if cols else df.copy()
    rename_map = {
        "表示相手先":"相手先",
        "借方中区分表示":"借方中区分",
        "借方小区分表示":"借方小区分",
        "貸方中区分表示":"貸方中区分",
        "貸方小区分表示":"貸方小区分",
    }
    show = show.rename(columns=rename_map)
    if "金額" in show.columns:
        show["金額"] = show["金額"].apply(fmt_yen)
    return show

def to_excel(summary, current_detail, history_detail, account, period_text, grain_mode, display_mode):
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    output = io.BytesIO()
    cond = pd.DataFrame({"項目":["表示モード","対象期間","勘定科目","集計粒度"], "内容":[display_mode, period_text, account, grain_mode]})

    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        summary.to_excel(writer, sheet_name="相手先別残高", index=False)
        current_detail.to_excel(writer, sheet_name="当月仕訳明細", index=False)
        history_detail.to_excel(writer, sheet_name="期首算出用履歴", index=False)
        cond.to_excel(writer, sheet_name="出力条件", index=False)

        header_fill = PatternFill(fill_type="solid", fgColor="D9E2F3")
        header_font = Font(bold=True)
        center = Alignment(horizontal="center", vertical="center")

        for sheet_name, df in {
            "相手先別残高": summary,
            "当月仕訳明細": current_detail,
            "期首算出用履歴": history_detail,
            "出力条件": cond,
        }.items():
            ws = writer.sheets[sheet_name]
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = center

            for col_idx, col_name in enumerate(df.columns, start=1):
                max_len = len(str(col_name))
                for row in range(2, ws.max_row + 1):
                    value = ws.cell(row=row, column=col_idx).value
                    if value is not None:
                        max_len = max(max_len, len(str(value)))
                    if col_name in ["期首残高","対象期間増加","対象期間減少","期末残高","増加","減少","金額"]:
                        ws.cell(row=row, column=col_idx).number_format = '#,##0'
                ws.column_dimensions[get_column_letter(col_idx)].width = max(12, min(32, max_len + 2))

    output.seek(0)
    return output.getvalue()

st.title("📘 会計内訳アプリ Pro 福祉の森完全対応版")
st.caption("前の機能を残したまま、中区分 / 小区分 の切替を追加した版です。")

with st.sidebar:
    with st.expander("データ読み込み", expanded=True):
        init_file = st.file_uploader("① 初期残高データ", type=["csv","xlsx","xlsm","xls"], key="init")
        journal_file = st.file_uploader("② 全期間仕訳マスタ（福祉の森CSV可）", type=["csv","xlsx","xlsm","xls"], key="journal")

    with st.expander("表示条件", expanded=True):
        bs_text = st.text_area("対象勘定科目（改行区切り）", value="\n".join(BS_DEFAULT), height=220)
        grain_mode = st.radio("集計粒度", ["中区分", "小区分"], index=0)
        hide_zero_partner = st.checkbox("期末残高ゼロの相手先を非表示", value=True)
        only_flags = st.checkbox("注意行のみ表示", value=False)
        partner_search = st.text_input("相手先検索")

    with st.expander("入力ルール", expanded=False):
        st.markdown("""
        - 上：初期残高データ  
        - 下：全期間仕訳マスタ or 福祉の森仕訳日記帳CSV  
        - 福祉の森の列名ゆらぎを吸収  
        - 中区分 / 小区分 の切替に対応  
        - 単月 / 期間累計 の切替に対応  
        - 2ヶ月以上動きがない残高は長期残存候補として⚠表示
        """)

if init_file and journal_file:
    try:
        kind1, df1 = parse_file(init_file)
        kind2, df2 = parse_file(journal_file)

        if kind1 == "journal" and kind2 == "initial":
            kind1, kind2 = kind2, kind1
            df1, df2 = df2, df1
            st.info("アップロード順を自動補正しました。")

        if kind1 != "initial" or kind2 != "journal":
            raise ValueError("上に初期残高、下に仕訳データをアップロードしてください。")

        initial_df = prepare_initial(df1)
        journal_df = prepare_journal(df2, grain_mode=grain_mode)

        accounts = choose_accounts(initial_df, journal_df, bs_text)
        months = month_options(initial_df, journal_df)
        if not months:
            raise ValueError("初期残高基準日より後の仕訳月がありません。")

        c1, c2, c3, c4 = st.columns([1,1,1,1])
        with c1:
            display_mode = st.radio("表示モード", ["単月", "期間累計"], horizontal=True)
        with c2:
            start_month = st.selectbox("開始月", months, index=len(months)-1)
        with c3:
            start_idx = months.index(start_month)
            if display_mode == "期間累計":
                end_month_candidates = months[start_idx:]
                end_month = st.selectbox("終了月", end_month_candidates, index=len(end_month_candidates)-1)
            else:
                end_month = start_month
                st.selectbox("終了月", [start_month], index=0, disabled=True)
        with c4:
            account = st.selectbox("勘定科目", accounts)

        period_text, period_label = make_period_labels(display_mode, start_month, end_month)
        summary, history_detail, current_detail, init_base, period_start, period_end = build_summary(
            initial_df, journal_df, account, start_month, end_month=end_month, mode=display_mode
        )

        if hide_zero_partner:
            summary = summary[summary["期末残高"] != 0].copy()
        if only_flags:
            summary = summary[(summary["長期残存候補"]) | (summary["マイナス残高"])].copy()
        if partner_search:
            summary = summary[summary["相手先"].astype(str).str.contains(partner_search, case=False, na=False)].copy()

        total_row = pd.Series({
            "期首残高": summary["期首残高"].astype(float).sum() if not summary.empty else 0,
            "対象期間増加": summary["対象期間増加"].astype(float).sum() if not summary.empty else 0,
            "対象期間減少": summary["対象期間減少"].astype(float).sum() if not summary.empty else 0,
            "期末残高": summary["期末残高"].astype(float).sum() if not summary.empty else 0,
        })

        st.markdown(
            f'<div class="info-box">基準日: {init_base.date()} / {period_label} / 勘定科目: {account} / 集計粒度: {grain_mode} / 表示モード: {display_mode}</div>',
            unsafe_allow_html=True
        )

        k1, k2, k3, k4 = st.columns(4)
        with k1: card("期首残高", total_row["期首残高"], "勘定科目全体の対象期間開始時点残高")
        with k2: card("増加", total_row["対象期間増加"], "勘定科目全体の対象期間借方計上")
        with k3: card("減少", total_row["対象期間減少"], "勘定科目全体の対象期間貸方計上")
        with k4: card("期末残高", total_row["期末残高"], "勘定科目全体の対象期間終了時点残高")

        detail_tab_name = "当月仕訳明細" if display_mode == "単月" else "対象期間仕訳明細"
        tab1, tab2, tab3, tab4 = st.tabs(["相手先別残高", detail_tab_name, "期首算出用履歴", "Excel出力"])

        with tab1:
            st.dataframe(style_summary(summary), use_container_width=True, hide_index=True)

        with tab2:
            st.dataframe(style_detail(current_detail), use_container_width=True, hide_index=True)
            st.caption("明細には1つの仕訳の借方科目・貸方科目の両方を表示しています。選択勘定科目の反対側科目も根拠確認のため表示されます。")

        with tab3:
            st.dataframe(style_detail(history_detail), use_container_width=True, hide_index=True)

        with tab4:
            excel_bytes = to_excel(summary, style_detail(current_detail), style_detail(history_detail), account, period_text, grain_mode, display_mode)
            st.download_button(
                "Excelダウンロード",
                data=excel_bytes,
                file_name=f"会計内訳_{period_text}_{account}_{grain_mode}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )

    except Exception as e:
        st.error(f"読み込みエラー：{e}")
else:
    st.markdown('<div class="subtle-box">左のサイドバーで、上に初期残高データ、下に福祉の森CSVを入れてください。</div>', unsafe_allow_html=True)
